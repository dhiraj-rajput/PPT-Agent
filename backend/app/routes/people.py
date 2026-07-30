"""
app/routes/people.py
-------------------------
People / Contacts CRM endpoints: search, filtering, manual entry, bulk manual
entry, and CSV / Excel import. Uses MySQL (people table) — see models.sql_models.Person.

Mirrors the structure of app/routes/companies.py so it plugs into the same
FastAPI app, auth, and DB-session patterns with no surprises.
"""

from __future__ import annotations

import codecs
import csv
import io
import logging
from datetime import datetime, date, timezone
from typing import Any, Dict, List, Optional

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from pydantic import BaseModel

from app.core.auth import get_current_user
from utils.db_client import get_db_session, _mysql_available
from models.sql_models import Person as SQL_Person
from sqlalchemy import select, insert, func, or_, and_

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/people", tags=["people"])

VALID_STATUSES = {"Pending", "Processing", "Completed", "Failed", "Duplicate"}
VALID_SOURCES = {"Apollo", "LinkedIn", "CSV Import", "Excel Import", "Manual Entry"}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _iso(dt: Any) -> Optional[str]:
    return dt.isoformat() if dt and hasattr(dt, "isoformat") else None


def _parse_date(val: Any):
    """Best-effort parse of a date coming from a form, CSV, or Excel cell."""
    if not val:
        return None
    if isinstance(val, (datetime, date)):
        return val
    val = str(val).strip()
    if not val:
        return None
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d-%m-%Y", "%d/%m/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(val, fmt)
        except ValueError:
            continue
    return None


def _format_person(p: SQL_Person) -> dict:
    if not p:
        return {}
    return {
        "id": str(p.id),
        "source": p.source or "Manual Entry",
        "status": p.status or "Pending",
        "organization_name": p.organization_name or "",
        "first_name": p.first_name or "",
        "last_name": p.last_name or "",
        "full_name": p.full_name or "",
        "title": p.title or "",
        "function_name": p.function_name or "",
        "seniority": p.seniority or "",
        "email": p.email or "",
        "email_status": p.email_status or "",
        "email_confidence": float(p.email_confidence) if p.email_confidence is not None else None,
        "phone": p.phone or "",
        "linkedin_url": p.linkedin_url or "",
        "city": p.city or "",
        "state": p.state or "",
        "country": p.country or "",
        "job_start_date": _iso(p.job_start_date),
        "createdAt": _iso(p.created_at),
        "updatedAt": _iso(p.updated_at),
    }


class PersonCreateBody(BaseModel):
    source: Optional[str] = "Manual Entry"
    status: Optional[str] = "Pending"
    organization_name: Optional[str] = ""
    first_name: Optional[str] = ""
    last_name: Optional[str] = ""
    full_name: Optional[str] = ""
    title: Optional[str] = ""
    function_name: Optional[str] = ""
    seniority: Optional[str] = ""
    email: Optional[str] = ""
    email_status: Optional[str] = ""
    email_confidence: Optional[float] = None
    phone: Optional[str] = ""
    linkedin_url: Optional[str] = ""
    city: Optional[str] = ""
    state: Optional[str] = ""
    country: Optional[str] = ""
    job_start_date: Optional[str] = ""


# ---------------------------------------------------------------------------
# List / Search / Filter / Paginate
# ---------------------------------------------------------------------------

@router.get("")
async def get_people(
    query: Optional[str] = None,
    status: Optional[str] = None,
    source: Optional[str] = None,
    country: Optional[str] = None,
    seniority: Optional[str] = None,
    page: int = 1,
    limit: int = 20,
    current_user: dict = Depends(get_current_user),
):
    """Query people from MySQL with unified search, filters, and pagination."""
    if not _mysql_available:
        raise HTTPException(status_code=500, detail="MySQL database is not connected.")

    try:
        filter_conditions = []

        if query and query.strip():
            q = query.strip()
            filter_conditions.append(or_(
                SQL_Person.full_name.ilike(f"%{q}%"),
                SQL_Person.first_name.ilike(f"%{q}%"),
                SQL_Person.last_name.ilike(f"%{q}%"),
                SQL_Person.email.ilike(f"%{q}%"),
                SQL_Person.organization_name.ilike(f"%{q}%"),
                SQL_Person.title.ilike(f"%{q}%"),
                SQL_Person.phone.ilike(f"%{q}%"),
            ))

        if status and status != "All":
            filter_conditions.append(SQL_Person.status == status)
        if source and source != "All":
            filter_conditions.append(SQL_Person.source == source)
        if country and country != "All":
            filter_conditions.append(SQL_Person.country == country)
        if seniority and seniority != "All":
            filter_conditions.append(SQL_Person.seniority == seniority)

        stmt = select(SQL_Person)
        if filter_conditions:
            stmt = stmt.where(and_(*filter_conditions))
        stmt = stmt.order_by(SQL_Person.created_at.desc())

        results: List[SQL_Person] = []
        total = 0
        source_options: List[str] = []
        country_options: List[str] = []

        async for db in get_db_session():
            count_stmt = select(func.count()).select_from(SQL_Person)
            if filter_conditions:
                count_stmt = count_stmt.where(and_(*filter_conditions))
            total = (await db.execute(count_stmt)).scalar() or 0

            skip = (page - 1) * limit
            res = await db.execute(stmt.offset(skip).limit(limit))
            results = res.scalars().all()

            src_res = await db.execute(select(SQL_Person.source).distinct())
            source_options = sorted([s for s in src_res.scalars().all() if s])

            country_res = await db.execute(select(SQL_Person.country).distinct())
            country_options = sorted([c for c in country_res.scalars().all() if c])

        return {
            "total": total,
            "page": page,
            "limit": limit,
            "people": [_format_person(p) for p in results],
            "source_options": source_options,
            "country_options": country_options,
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ---------------------------------------------------------------------------
# Manual add (single record)
# ---------------------------------------------------------------------------

@router.post("")
async def add_person(
    person_data: PersonCreateBody,
    current_user: dict = Depends(get_current_user),
):
    """Add a single person record manually to MySQL."""
    if not _mysql_available:
        raise HTTPException(status_code=500, detail="MySQL database is unavailable.")

    full_name = (person_data.full_name or "").strip()
    if not full_name:
        full_name = f"{(person_data.first_name or '').strip()} {(person_data.last_name or '').strip()}".strip()
    if not full_name:
        raise HTTPException(status_code=400, detail="Full name (or first/last name) is required")

    try:
        async for db in get_db_session():
            stmt_ins = insert(SQL_Person).values(
                source=person_data.source or "Manual Entry",
                status=person_data.status or "Pending",
                organization_name=person_data.organization_name or "",
                first_name=person_data.first_name or "",
                last_name=person_data.last_name or "",
                full_name=full_name,
                title=person_data.title or "",
                function_name=person_data.function_name or "",
                seniority=person_data.seniority or "",
                email=person_data.email or "",
                email_status=person_data.email_status or "",
                email_confidence=person_data.email_confidence,
                phone=person_data.phone or "",
                linkedin_url=person_data.linkedin_url or "",
                city=person_data.city or "",
                state=person_data.state or "",
                country=person_data.country or "",
                job_start_date=_parse_date(person_data.job_start_date),
                created_at=datetime.utcnow(),
                updated_at=datetime.utcnow(),
            )
            await db.execute(stmt_ins)
            await db.commit()
            return {"status": "success", "message": "Person added successfully"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Database error: {e}")


# ---------------------------------------------------------------------------
# Bulk manual entry / small JSON import
# ---------------------------------------------------------------------------

@router.post("/import")
async def import_people(payload: dict, current_user: dict = Depends(get_current_user)):
    """
    Bulk import people from a JSON array of records (used by both the
    "bulk manual entry" grid and small programmatic imports). For large
    CSV/Excel files use /import/file instead.
    """
    items = payload.get("data") or []
    if isinstance(items, str):
        import json
        items = json.loads(items)

    imported_count = 0
    skipped = 0
    try:
        async for db in get_db_session():
            for item in items:
                try:
                    validated = PersonCreateBody(**item)
                except Exception:
                    skipped += 1
                    continue

                full_name = (validated.full_name or "").strip()
                if not full_name:
                    full_name = f"{(validated.first_name or '').strip()} {(validated.last_name or '').strip()}".strip()
                if not full_name:
                    skipped += 1
                    continue

                await db.execute(insert(SQL_Person).values(
                    source=validated.source or "Manual Entry",
                    status=validated.status or "Pending",
                    organization_name=validated.organization_name or "",
                    first_name=validated.first_name or "",
                    last_name=validated.last_name or "",
                    full_name=full_name,
                    title=validated.title or "",
                    function_name=validated.function_name or "",
                    seniority=validated.seniority or "",
                    email=validated.email or "",
                    email_status=validated.email_status or "",
                    email_confidence=validated.email_confidence,
                    phone=validated.phone or "",
                    linkedin_url=validated.linkedin_url or "",
                    city=validated.city or "",
                    state=validated.state or "",
                    country=validated.country or "",
                    job_start_date=_parse_date(validated.job_start_date),
                    created_at=datetime.utcnow(),
                    updated_at=datetime.utcnow(),
                ))
                imported_count += 1
            await db.commit()
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Import failed: {e}")

    return {"status": "success", "count": imported_count, "skipped": skipped}


# ---------------------------------------------------------------------------
# CSV / Excel file import (streamed, chunked)
# ---------------------------------------------------------------------------

_ROW_ALIASES = {
    "full_name": ["full_name", "Full Name", "Full_Name", "name", "Name"],
    "first_name": ["first_name", "First Name", "First_Name"],
    "last_name": ["last_name", "Last Name", "Last_Name"],
    "organization_name": ["organization_name", "Organization", "Company", "organization", "company_name", "Organization Name"],
    "title": ["title", "Title", "Job Title"],
    "function_name": ["function_name", "Function", "Department"],
    "seniority": ["seniority", "Seniority"],
    "email": ["email", "Email", "Email Address"],
    "email_status": ["email_status", "Email Status"],
    "email_confidence": ["email_confidence", "Email Confidence"],
    "phone": ["phone", "Phone", "Phone Number"],
    "linkedin_url": ["linkedin_url", "LinkedIn", "LinkedIn URL", "Linkedin Url"],
    "city": ["city", "City"],
    "state": ["state", "State"],
    "country": ["country", "Country"],
    "job_start_date": ["job_start_date", "Job Start Date", "Start Date"],
    "source": ["source", "Source"],
    "status": ["status", "Status"],
}


def _get_field(row: Dict[str, Any], field: str) -> str:
    for alias in _ROW_ALIASES.get(field, [field]):
        if alias in row and row[alias] not in (None, ""):
            return str(row[alias]).strip()
    return ""


def _row_to_person_dict(row: Dict[str, Any], default_source: str) -> Optional[dict]:
    full_name = _get_field(row, "full_name")
    first_name = _get_field(row, "first_name")
    last_name = _get_field(row, "last_name")
    if not full_name:
        full_name = f"{first_name} {last_name}".strip()
    if not full_name:
        return None

    status_val = _get_field(row, "status") or "Pending"
    if status_val not in VALID_STATUSES:
        status_val = "Pending"

    conf_raw = _get_field(row, "email_confidence")
    try:
        confidence = round(float(conf_raw), 2) if conf_raw else None
    except ValueError:
        confidence = None

    return {
        "source": _get_field(row, "source") or default_source,
        "status": status_val,
        "organization_name": _get_field(row, "organization_name"),
        "first_name": first_name,
        "last_name": last_name,
        "full_name": full_name,
        "title": _get_field(row, "title"),
        "function_name": _get_field(row, "function_name"),
        "seniority": _get_field(row, "seniority"),
        "email": _get_field(row, "email"),
        "email_status": _get_field(row, "email_status"),
        "email_confidence": confidence,
        "phone": _get_field(row, "phone"),
        "linkedin_url": _get_field(row, "linkedin_url"),
        "city": _get_field(row, "city"),
        "state": _get_field(row, "state"),
        "country": _get_field(row, "country"),
        "job_start_date": _parse_date(_get_field(row, "job_start_date")),
        "created_at": datetime.utcnow(),
        "updated_at": datetime.utcnow(),
    }


async def _flush_chunk(chunk: List[dict]) -> int:
    if not chunk:
        return 0
    async for db in get_db_session():
        await db.execute(insert(SQL_Person).values(chunk))
        await db.commit()
        return len(chunk)
    return 0


async def _process_csv_stream(text_stream, default_source: str = "CSV Import") -> int:
    reader = csv.DictReader(text_stream)
    imported_count = 0
    chunk: List[dict] = []
    CHUNK = 500

    for row in reader:
        if not isinstance(row, dict):
            continue
        doc = _row_to_person_dict(row, default_source)
        if not doc:
            continue
        chunk.append(doc)
        if len(chunk) >= CHUNK:
            imported_count += await _flush_chunk(chunk)
            chunk = []

    if chunk:
        imported_count += await _flush_chunk(chunk)

    return imported_count


async def _process_xlsx_bytes(file_bytes: bytes, default_source: str = "Excel Import") -> int:
    """Read the first worksheet of an .xlsx/.xls workbook and bulk-insert rows."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl is required to import Excel files.")

    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active

    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = [str(h).strip() if h is not None else "" for h in next(rows_iter)]
    except StopIteration:
        return 0

    imported_count = 0
    chunk: List[dict] = []
    CHUNK = 500

    for values in rows_iter:
        if values is None or all(v is None for v in values):
            continue
        row = {header[i]: values[i] for i in range(min(len(header), len(values)))}
        doc = _row_to_person_dict(row, default_source)
        if not doc:
            continue
        chunk.append(doc)
        if len(chunk) >= CHUNK:
            imported_count += await _flush_chunk(chunk)
            chunk = []

    if chunk:
        imported_count += await _flush_chunk(chunk)

    return imported_count


@router.post("/import/file")
async def import_people_file(
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user),
):
    """
    Stream-import people from a CSV, XLS, or XLSX file upload into MySQL.
    Recognised header aliases are listed in _ROW_ALIASES above — anything
    else is ignored, so exports from Apollo/LinkedIn/HubSpot etc. work
    without pre-cleaning the file first.
    """
    filename = (file.filename or "").lower()
    if not (filename.endswith(".csv") or filename.endswith(".xlsx") or filename.endswith(".xls")):
        raise HTTPException(status_code=400, detail="Please upload a .csv, .xlsx, or .xls file.")

    try:
        if filename.endswith(".csv"):
            text_stream = codecs.iterdecode(file.file, "utf-8", errors="replace")
            imported_count = await _process_csv_stream(text_stream)
        else:
            file_bytes = await file.read()
            imported_count = await _process_xlsx_bytes(file_bytes)
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"People file import failed: {e}", exc_info=True)
        raise HTTPException(status_code=400, detail=f"File processing failed: {str(e)}")
    finally:
        await file.close()

    return {"status": "success", "count": imported_count}


# ---------------------------------------------------------------------------
# Detail
# ---------------------------------------------------------------------------

@router.get("/{person_id}")
async def get_person_detail(
    person_id: str,
    current_user: dict = Depends(get_current_user),
):
    """Retrieve full details of a specific person from MySQL."""
    if not _mysql_available:
        raise HTTPException(status_code=500, detail="MySQL database is unavailable.")

    try:
        pid = int(person_id)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid person id")

    try:
        async for db in get_db_session():
            stmt = select(SQL_Person).where(SQL_Person.id == pid)
            res = await db.execute(stmt)
            person = res.scalar_one_or_none()
            if not person:
                raise HTTPException(status_code=404, detail="Person not found")
            return _format_person(person)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
