"""
app/core/company_catalog.py
----------------------------
Loads and indexes the authoritative OrbitAvanya services and add-ons catalog.

Read order:
  1. MongoDB collection `company_catalog` (document _id="orbitavanya") — the
     live source of truth once `scripts/import_company_catalog.py` has been run.
  2. private/OrbitAvanya_Services_ADD.xlsx, if present, as a one-time seed source.
  3. A tiny hardcoded stub, only so the pipeline never hard-fails.
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Any, Dict, List, Optional

logger = logging.getLogger(__name__)

PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
EXCEL_PATH = PROJECT_ROOT / "private" / "OrbitAvanya_Services_ADD.xlsx"
MONGO_COLLECTION = "company_catalog"
MONGO_DOC_ID = "orbitavanya"

_cached_catalog: Optional[Dict[str, Any]] = None


def _load_from_mongo() -> Optional[Dict[str, Any]]:
    try:
        from utils.db_client import get_collection
        doc = get_collection(MONGO_COLLECTION).find_one({"_id": MONGO_DOC_ID})
        if doc and doc.get("services"):
            categories: Dict[str, List[Dict[str, Any]]] = {}
            for item in doc["services"]:
                categories.setdefault(item.get("category", "General"), []).append(item)
            logger.info(
                f"[CompanyCatalog] Loaded {len(doc['services'])} services and "
                f"{len(doc.get('addons', []))} add-ons from MongoDB."
            )
            return {
                "services": doc["services"],
                "addons": doc.get("addons", []),
                "categories": categories,
            }
    except Exception as exc:
        logger.warning(f"[CompanyCatalog] MongoDB catalog read failed, falling back: {exc}")
    return None


def _parse_excel_workbook(path: Path) -> Optional[Dict[str, Any]]:
    """Parses both sheets ('Services' and 'Premium Enterprise Add-ons') into the
    same shape as _load_from_mongo(). Shared by the fallback path here and by
    scripts/import_company_catalog.py when seeding MongoDB."""
    if not path.exists():
        return None
    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(path), data_only=True)

        services: List[Dict[str, Any]] = []
        addons: List[Dict[str, Any]] = []
        categories: Dict[str, List[Dict[str, Any]]] = {}

        if "Services" in wb.sheetnames:
            ws = wb["Services"]
            for row in list(ws.iter_rows(min_row=2, values_only=True)):
                if not row or not any(row):
                    continue
                naics = str(row[1]).strip() if len(row) > 1 and row[1] else "541511"
                category = str(row[2]).strip() if len(row) > 2 and row[2] else "General"
                service_name = str(row[3]).strip() if len(row) > 3 and row[3] else ""
                description = str(row[4]).strip() if len(row) > 4 and row[4] else ""
                if not service_name:
                    continue
                item = {
                    "id": row[0],
                    "naics_code": naics,
                    "category": category,
                    "service_name": service_name,
                    "description": description,
                }
                services.append(item)
                categories.setdefault(category, []).append(item)

        if "Premium Enterprise Add-ons" in wb.sheetnames:
            ws_add = wb["Premium Enterprise Add-ons"]
            for row in list(ws_add.iter_rows(min_row=2, values_only=True)):
                if not row or not any(row):
                    continue
                name = str(row[0]).strip() if len(row) > 0 and row[0] else ""
                price = str(row[1]).strip() if len(row) > 1 and row[1] else "Custom Pricing"
                if name and name.lower() not in ("service", "add-on", "support plan"):
                    addons.append({"service_name": name, "price": price})

        logger.info(f"[CompanyCatalog] Parsed {len(services)} services and {len(addons)} add-ons from Excel.")
        return {"services": services, "addons": addons, "categories": categories}
    except Exception as exc:
        logger.warning(f"[CompanyCatalog] Failed to parse Excel catalog {path}: {exc}")
        return None


def load_services_catalog(force_refresh: bool = False) -> Dict[str, Any]:
    global _cached_catalog
    if _cached_catalog is not None and not force_refresh:
        return _cached_catalog

    catalog = _load_from_mongo() or _parse_excel_workbook(EXCEL_PATH)

    if not catalog or not catalog.get("services"):
        # Last-resort stub so the pipeline never hard-fails — this should not be
        # reached once the MongoDB catalog is seeded via the import script.
        stub_services = [
            {"id": 1, "naics_code": "541511", "category": "AI Solutions", "service_name": "AI Chatbot & RAG", "description": "Enterprise RAG & Conversational AI"},
            {"id": 2, "naics_code": "541511", "category": "Custom Software", "service_name": "ERP & CRM Development", "description": "Enterprise ERP, CRM & HRMS Systems"},
            {"id": 3, "naics_code": "541511", "category": "Web Development", "service_name": "Government Portal", "description": "Secure Citizen Portals & CMS"},
            {"id": 4, "naics_code": "541511", "category": "Mobile Apps", "service_name": "Flutter & Native Apps", "description": "iOS & Android Cross-Platform Applications"},
        ]
        categories: Dict[str, List[Dict[str, Any]]] = {}
        for item in stub_services:
            categories.setdefault(item["category"], []).append(item)
        catalog = {"services": stub_services, "addons": [], "categories": categories}
        logger.warning("[CompanyCatalog] No MongoDB or Excel catalog found — using minimal stub. "
                       "Run scripts/import_company_catalog.py to seed the real catalog.")

    _cached_catalog = catalog
    return _cached_catalog


def get_catalog_products() -> List[Dict[str, Any]]:
    """Returns a list of structured product profiles matching the format expected by generators."""
    catalog = load_services_catalog()
    products: List[Dict[str, Any]] = []

    for cat_name, cat_items in catalog["categories"].items():
        features = [item["service_name"] for item in cat_items]
        descriptions = " ".join(item["description"] for item in cat_items if item["description"])
        
        products.append({
            "product_name": f"OrbitAvanya {cat_name}",
            "company_name": "OrbitAvanya Tech LLP",
            "industry_domain": cat_name,
            "about_text": f"OrbitAvanya's {cat_name} suite provides: {descriptions[:300]}...",
            "key_features": features,
            "technology_stack": {
                "frontend": ["React", "Next.js", "Flutter"],
                "backend": ["Python", "FastAPI", "Node.js"],
                "database": ["PostgreSQL", "MongoDB"],
                "cloud": ["AWS", "Azure", "Docker"]
            },
            "security_and_compliance": ["ISO 27001 Ready", "SOC 2 Type II", "Role-Based Access Control", "FIPS 140 Encryption"],
            "pricing_model": "Custom Enterprise SLA / Fixed Milestone"
        })

    return products


def get_catalog_summary_for_prompt(max_services: int = 40, max_addons: int = 20) -> str:
    """Human-readable catalog summary formatted for direct inclusion in an LLM
    prompt — grouped by category so the AI can pick genuinely relevant services
    for the solicitation at hand instead of inventing capabilities."""
    catalog = load_services_catalog()
    lines: List[str] = []
    count = 0
    for category, items in catalog["categories"].items():
        if count >= max_services:
            break
        lines.append(f"{category}:")
        for item in items:
            if count >= max_services:
                break
            lines.append(f"  - {item['service_name']} (NAICS {item['naics_code']}): {item['description']}")
            count += 1

    if catalog.get("addons"):
        lines.append("\nPremium / Enterprise Add-ons (real starting prices — use these as pricing anchors, never invent numbers):")
        for addon in catalog["addons"][:max_addons]:
            lines.append(f"  - {addon['service_name']}: {addon['price']}")

    return "\n".join(lines) if lines else "No catalog data available."
